"""
SET Financial Statement Scraper
================================
Fetches financial data from SET.or.th using the internal API endpoints.

Data Flow:
1. Init session -> GET /th/home to obtain Incapsula cookies
2. GET /api/set/news/{symbol}/list -> find news with tag='financial-statement'
3. GET /api/set/news/{id}/detail -> get downloadUrl (.zip)
4. Download ZIP from weblink.set.or.th -> extract FINANCIAL_STATEMENTS.XLSX
5. GET /api/set/stock/{symbol}/company-highlight/financial-data -> annual summary
6. Parse XLSX for detailed line items
"""

import io
import os
import time
import json
import zipfile
import tempfile
import requests
import openpyxl
import pandas as pd
from typing import Optional
from pathlib import Path


# ---------------------------------------------------------------------------
# Cache directory
# ---------------------------------------------------------------------------
CACHE_DIR = Path(tempfile.gettempdir()) / "set_financial_cache"
CACHE_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Thai period description parser (module-level for use in both scraper & data)
# ---------------------------------------------------------------------------
def _parse_thai_period_static(desc: str) -> dict:
    """
    Parse Thai period description to extract months_count and other info.
    E.g. "สำหรับงวดหกเดือนสิ้นสุดวันที่ 30 มิถุนายน 2568" → months_count=6
    """
    import re
    result = {"month": 0, "year_ce": 0, "months_count": 0}
    if not desc:
        return result

    _thai_months = {
        "มกราคม": 1, "กุมภาพันธ์": 2, "มีนาคม": 3, "เมษายน": 4,
        "พฤษภาคม": 5, "มิถุนายน": 6, "กรกฎาคม": 7, "สิงหาคม": 8,
        "กันยายน": 9, "ตุลาคม": 10, "พฤศจิกายน": 11, "ธันวาคม": 12,
    }
    for th, num in _thai_months.items():
        if th in desc:
            result["month"] = num
            break

    year_match = re.search(r"(\d{4})", desc)
    if year_match:
        y = int(year_match.group(1))
        result["year_ce"] = y - 543 if y > 2400 else y

    month_words = {
        "สามเดือน": 3, "หกเดือน": 6, "เก้าเดือน": 9,
        "สิบสองเดือน": 12, "หนึ่งปี": 12,
        # English fallbacks
        "three-month": 3, "three month": 3,
        "six-month": 6, "six month": 6,
        "nine-month": 9, "nine month": 9,
        "twelve-month": 12, "twelve month": 12,
    }
    for w, n in month_words.items():
        if w in desc.lower():
            result["months_count"] = n
            break

    return result


class SETScraper:
    """Scraper for SET.or.th financial statements via internal APIs."""

    BASE_URL = "https://www.set.or.th"
    API_BASE = "https://www.set.or.th/api/set"

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9,th;q=0.8",
            "Referer": (
                "https://www.set.or.th/th/market/product/stock/"
                "quote/AOT/financial-statement/company-highlights"
            ),
        })
        self._session_ready = False

    # ------------------------------------------------------------------
    # Session management
    # ------------------------------------------------------------------
    def _ensure_session(self):
        """Visit the main page once to obtain Incapsula cookies."""
        if self._session_ready:
            return
        try:
            self.session.get(f"{self.BASE_URL}/th/home", timeout=15)
            self._session_ready = True
        except Exception:
            pass

    def _api_get(self, endpoint: str, params: dict | None = None):
        """GET from the SET internal API with automatic retry."""
        self._ensure_session()
        url = f"{self.API_BASE}/{endpoint}"
        for attempt in range(3):
            try:
                r = self.session.get(url, params=params, timeout=20)
                if r.status_code == 200:
                    return r.json()
                if r.status_code == 403:
                    self._session_ready = False
                    self._ensure_session()
                    time.sleep(1)
            except Exception:
                if attempt < 2:
                    time.sleep(1)
        return None

    # ------------------------------------------------------------------
    # High-level data fetchers
    # ------------------------------------------------------------------
    def get_company_highlight_financial(self, symbol: str) -> list[dict] | None:
        """
        Financial summary via company-highlight/financial-data.
        Returns list of dicts (all available periods), newest first.
        Values are in *thousands* of THB.

        The API returns both annual (Q9/YE/Q4) and quarterly records.
        Use filter_annual() / filter_quarterly() helpers to separate them.
        """
        data = self._api_get(
            f"stock/{symbol.upper()}/company-highlight/financial-data",
            {"lang": "en"},
        )
        if isinstance(data, list) and data:
            data.sort(key=lambda x: (x.get("year", 0), x.get("quarter", "")), reverse=True)
            return data
        return None

    def get_quarterly_highlight_financial(self, symbol: str) -> list[dict] | None:
        """
        Quarterly financial summary via company-highlight/financial-data
        with quarter=Q parameter.
        Returns list of dicts for quarterly data, newest first.
        Values are in *thousands* of THB.
        """
        data = self._api_get(
            f"stock/{symbol.upper()}/company-highlight/financial-data",
            {"lang": "en", "quarter": "Q"},
        )
        if isinstance(data, list) and data:
            data.sort(key=lambda x: (x.get("year", 0), x.get("quarter", "")), reverse=True)
            return data
        return None

    def get_factsheet_submissions(self, symbol: str) -> list[dict] | None:
        """
        Get list of all quarterly financial statement submissions.
        Returns list with submitDate, quarter (Q1/Q2/Q3/YE), year.
        """
        data = self._api_get(
            f"factsheet/{symbol.upper()}/financialstatement-submission",
            {"lang": "en"},
        )
        if isinstance(data, list) and data:
            return data
        return None

    def get_key_financial_data(self, symbol: str) -> dict | None:
        """Latest quarter key financial data."""
        return self._api_get(
            f"stock/{symbol.upper()}/key-financial-data",
            {"lang": "en"},
        )

    # ------------------------------------------------------------------
    # Factsheet financial statement API (multi-period structured data)
    # ------------------------------------------------------------------
    def get_factsheet_financialstatement(
        self, symbol: str, account_type: str, lang: str = "en"
    ) -> list[dict] | None:
        """
        Get multi-period financial statement line items from the factsheet API.

        account_type: "income_statement", "balance_sheet", or "cash_flow"
        Returns list of period objects, each with 'accountList' containing
        line items.  Typically 4-5 periods (mix of quarterly and annual).

        Values use 'divider' field (usually 1000 = thousands THB, except
        EPS which has divider=1).
        """
        data = self._api_get(
            f"factsheet/{symbol.upper()}/financialstatement",
            {"accountType": account_type, "lang": lang},
        )
        if isinstance(data, list) and data:
            return data
        return None

    def get_factsheet_financial_ratio(self, symbol: str) -> list[dict] | None:
        """
        Get multi-period financial ratios from factsheet API.
        Returns list of period objects with ratio items.
        Typically 4 periods.
        """
        data = self._api_get(
            f"factsheet/{symbol.upper()}/financial-ratio",
            {"lang": "en"},
        )
        if isinstance(data, list) and data:
            return data
        return None

    def get_factsheet_financial_growth(self, symbol: str) -> list[dict] | None:
        """Get multi-period growth metrics from factsheet API."""
        data = self._api_get(
            f"factsheet/{symbol.upper()}/financial-growth",
            {"lang": "en"},
        )
        if isinstance(data, list) and data:
            return data
        return None

    def get_company_profile(self, symbol: str) -> dict | None:
        """Company profile (name, sector, market, etc.)."""
        return self._api_get(
            f"stock/{symbol.upper()}/profile",
            {"lang": "en"},
        )

    def get_company_profile_factsheet(self, symbol: str) -> dict | None:
        """Company profile via factsheet endpoint."""
        return self._api_get(
            f"factsheet/{symbol.upper()}/profile",
            {"lang": "en"},
        )

    def get_stock_list(self) -> list | None:
        """Get full list of SET stocks."""
        return self._api_get("stock/list", {"lang": "en"})

    # ------------------------------------------------------------------
    # News-based financial statement ZIP fetcher
    # ------------------------------------------------------------------
    def get_all_fs_news(self, symbol: str, years_back: int = 5) -> list[dict]:
        """
        Get ALL financial-statement news for a symbol going back N years.

        Uses news/search API with fromDate/toDate.  The API allows max ~4
        years per call, so we split into multiple date-range chunks.

        Returns list of news items (tag='financial-statement') with basic
        info (id, headline, datetime).  Call get_fs_news_detail(id) to
        get the downloadUrl.
        """
        from datetime import datetime, timedelta

        symbol = symbol.upper()
        all_items: list[dict] = []
        seen_ids: set[str] = set()

        now = datetime.now()
        end_date = now
        chunk_days = 180  # 6 months per chunk — 1-year chunks sometimes return None for older data

        start_limit = now - timedelta(days=years_back * 365 + 60)

        while end_date > start_limit:
            start_date = end_date - timedelta(days=chunk_days)
            if start_date < start_limit:
                start_date = start_limit

            from_str = start_date.strftime("%d/%m/%Y")
            to_str = end_date.strftime("%d/%m/%Y")

            data = self._api_get("news/search", {
                "symbol": symbol,
                "fromDate": from_str,
                "toDate": to_str,
                "pageSize": 500,
                "lang": "th",
            })

            if data and "newsInfoList" in data:
                for item in data["newsInfoList"]:
                    if item.get("tag") == "financial-statement":
                        nid = str(item.get("id", ""))
                        if nid and nid not in seen_ids:
                            seen_ids.add(nid)
                            all_items.append(item)

            # Move window back
            end_date = start_date - timedelta(days=1)

        # Sort by date descending (newest first)
        all_items.sort(key=lambda x: x.get("datetime", ""), reverse=True)
        return all_items

    def get_fs_news_detail(self, news_id: str) -> dict | None:
        """Get detail for a single news item, including downloadUrl."""
        return self._api_get(f"news/{news_id}/detail", {"lang": "th"})

    @staticmethod
    def _parse_fs_headline(headline: str) -> tuple[str, int]:
        """
        Parse quarter and CE year from financial statement news headline.

        Examples:
          "งบการเงิน ประจำปี 2568 (ตรวจสอบแล้ว)"  → ("Q9", 2025)
          "งบการเงิน ไตรมาสที่ 3/2568 (สอบทานแล้ว)" → ("Q3", 2025)
          "งบการเงิน ไตรมาสที่ 1/2567 (สอบทานแล้ว)" → ("Q1", 2024)

        Returns (quarter, year_ce).  quarter is "Q1"/"Q2"/"Q3"/"Q9".
        """
        import re
        quarter = "Q9"  # default = annual
        year_ce = 0

        # Quarter: ไตรมาสที่ 1/2568, ไตรมาสที่ 2/2567, etc.
        qm = re.search(r"ไตรมาสที่\s*(\d)", headline)
        if qm:
            quarter = f"Q{qm.group(1)}"

        # Year: 4-digit Thai year (e.g. 2568 → CE 2025)
        ym = re.search(r"(\d{4})", headline)
        if ym:
            y = int(ym.group(1))
            year_ce = y - 543 if y > 2400 else y

        return quarter, year_ce

    def get_latest_fs_zip_url(self, symbol: str) -> dict | None:
        """
        Get the latest full financial statement ZIP download info.
        Uses the latest-full-financialstatement endpoint (fastest).
        """
        data = self._api_get(
            f"stock/{symbol.upper()}/financialstatement/latest-full-financialstatement",
            {"lang": "th"},
        )
        return data

    # ------------------------------------------------------------------
    # ZIP download & extraction
    # ------------------------------------------------------------------
    def download_zip(self, url: str) -> bytes | None:
        """Download a ZIP file from weblink.set.or.th."""
        try:
            r = self.session.get(url, timeout=30)
            if r.status_code == 200 and len(r.content) > 100:
                return r.content
        except Exception:
            pass
        return None

    def extract_xlsx_from_zip(self, zip_bytes: bytes) -> bytes | None:
        """
        Extract FINANCIAL_STATEMENTS from ZIP bytes.

        Handles both .XLSX and .XLS extensions.  Some SET filings use
        a .XLS extension but the file is actually in XLSX (OOXML) format,
        so we first try openpyxl regardless of extension.  If that fails
        and the file is truly old-format .XLS, we convert via xlrd.
        """
        try:
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                # Find the best candidate (FINANCIAL_STATEMENTS.XLSX or .XLS)
                candidates: list[str] = []
                for name in zf.namelist():
                    upper = name.upper()
                    if upper.endswith(".XLSX") or upper.endswith(".XLS"):
                        if "FINANCIAL" in upper:
                            candidates.insert(0, name)   # prioritise
                        else:
                            candidates.append(name)

                for name in candidates:
                    raw = zf.read(name)
                    # Try as XLSX first (works even for .XLS-named OOXML files)
                    try:
                        openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                        return raw                        # valid XLSX
                    except Exception:
                        pass
                    # Fall back: convert genuine old-format XLS → XLSX
                    converted = self._xls_to_xlsx(raw)
                    if converted:
                        return converted
        except Exception:
            pass
        return None

    @staticmethod
    def _xls_to_xlsx(xls_bytes: bytes) -> bytes | None:
        """Convert old .XLS bytes to .XLSX bytes using pandas + xlrd."""
        try:
            xls_io = io.BytesIO(xls_bytes)
            sheets = pd.read_excel(xls_io, sheet_name=None, header=None, engine="xlrd")
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            return out.getvalue()
        except Exception:
            return None

    # ------------------------------------------------------------------
    # XLSX Parsing
    # ------------------------------------------------------------------
    @staticmethod
    def _classify_sheet(ws) -> tuple[str, str]:
        """
        Classify a worksheet by examining its first ~8 rows of content.

        Returns (sheet_type, period_type) where:
          sheet_type: "bs", "income", "comprehensive", "cashflow",
                      "equity_changes", "unknown"
          period_type: "standalone", "cumulative", "annual", "point_in_time", ""

        Works regardless of sheet naming conventions.
        """
        texts: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row),
                                values_only=True):
            for c in row:
                if isinstance(c, str) and c.strip():
                    texts.append(c.strip())

        joined = " ".join(texts)

        # --- Determine sheet_type ---
        sheet_type = "unknown"
        if (
            "งบฐานะการเงิน" in joined
            or "งบแสดงฐานะการเงิน" in joined
            or "Statement of Financial Position" in joined
        ):
            sheet_type = "bs"
        elif (
            "งบกระแสเงินสด" in joined
            or "Statement of cash flows" in joined
            or "Statements of cash flows" in joined
            or "Statement of Cash Flows" in joined
            or "Statements of Cash Flows" in joined
            or "Net cash provided by operating activities" in joined
            or "Net cash flows from operating activities" in joined
        ):
            sheet_type = "cashflow"
        elif "การเปลี่ยนแปลงส่วนของผู้ถือหุ้น" in joined \
                or "Changes in Equity" in joined \
                or "Changes in Shareholders" in joined:
            sheet_type = "equity_changes"
        elif "กำไรขาดทุนเบ็ดเสร็จ" in joined \
                or "Comprehensive Income" in joined \
                or "comprehensive income" in joined:
            # Combined income + comprehensive (KBANK, GFPT style) OR
            # standalone comprehensive income sheet.
            # If it also contains "กำไรขาดทุน" (without เบ็ดเสร็จ alone),
            # treat as income (it includes both).
            sheet_type = "income"  # comprehensive includes income data
        elif "กำไรขาดทุน" in joined \
                or "Income Statement" in joined \
                or "Profit or Loss" in joined \
                or "Profit and Loss" in joined \
                or "Statement of Income" in joined:
            sheet_type = "income"

        # --- Determine period_type ---
        period_type = ""
        if sheet_type == "bs":
            period_type = "point_in_time"
        elif "สามเดือน" in joined or "three-month" in joined.lower() \
                or "three month" in joined.lower() \
                or "(3ด)" in joined or "(3m)" in joined.lower() \
                or "3 เดือน" in joined:
            period_type = "standalone"
        elif "หกเดือน" in joined or "เก้าเดือน" in joined \
                or "six-month" in joined.lower() or "nine-month" in joined.lower() \
                or "(6ด)" in joined or "(9ด)" in joined \
                or "(6m)" in joined.lower() or "(9m)" in joined.lower() \
                or "6 เดือน" in joined or "9 เดือน" in joined:
            period_type = "cumulative"
        elif "สิบสองเดือน" in joined or "twelve" in joined.lower() \
                or "สำหรับปี" in joined or "For the year" in joined:
            period_type = "annual"

        return sheet_type, period_type

    def parse_financial_xlsx(self, xlsx_bytes: bytes) -> dict:
        """
        Parse FINANCIAL_STATEMENTS.XLSX into structured data.
        Returns dict with keys: balance_sheet, income_statement, cashflow, metadata.

        Uses **content-based detection** to classify sheets — works regardless
        of sheet naming conventions (numeric, abbreviated, Thai, English).

        For income statement sheets that contain MULTIPLE sections (e.g.
        Q2 ZIPs with standalone 3-month + cumulative 6-month in one sheet),
        only the first standalone section is kept in 'income_statement'.
        """
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        result = {
            "sheets": wb.sheetnames,
            "balance_sheet": {},
            "income_statement": {},
            "income_statement_cumulative": {},
            "cashflow": {},
        }

        # --- Phase 1: classify every sheet by content -----------------------
        bs_sheets: list = []
        income_standalone: list = []   # standalone 3-month income sheets
        income_annual: list = []       # full-year (12-month) income sheets
        income_cumulative: list = []   # cumulative (6/9-month) income sheets
        income_unknown: list = []      # income sheets with unclear period
        cashflow_sheets: list = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            stype, ptype = self._classify_sheet(ws)

            if stype == "bs":
                bs_sheets.append(ws)
            elif stype == "income":
                if ptype == "standalone":
                    income_standalone.append((sheet_name, ws))
                elif ptype == "cumulative":
                    income_cumulative.append((sheet_name, ws))
                elif ptype == "annual":
                    # Keep annual sheets separate — they have full-year values
                    # in whatever unit the annual report uses (may differ from
                    # the quarterly sheets in the same ZIP).
                    income_annual.append((sheet_name, ws))
                else:
                    income_unknown.append((sheet_name, ws))
            elif stype == "cashflow":
                cashflow_sheets.append(ws)
            # equity_changes and unknown are ignored

        # --- Phase 1b: detect combined PL+CF sheets -------------------------
        # Some companies (e.g. INET) put income statement AND cashflow into
        # ONE sheet ("FS-PL,CF").  _classify_sheet() sees the income header
        # first and classifies it as "income", missing the cashflow section.
        # If no dedicated cashflow sheet was found, scan income sheets for
        # cashflow content and add them to cashflow_sheets as well.
        # We record the CF start row so _parse_sheet can begin there (to get
        # the correct period_description metadata for cashflow).
        _cf_start_row: int | None = None  # 1-indexed row where CF section begins
        if not cashflow_sheets:
            _cf_markers = ("งบกระแสเงินสด", "Statement of cash flows",
                           "Statements of cash flows", "Statement of Cash Flows",
                           "Statements of Cash Flows")
            all_income_ws = (
                [(sn, ws) for sn, ws in income_standalone]
                + [(sn, ws) for sn, ws in income_cumulative]
                + [(sn, ws) for sn, ws in income_annual]
                + [(sn, ws) for sn, ws in income_unknown]
            )
            for _sn, _ws in all_income_ws:
                # Scan rows beyond the first 15 (skip income header area)
                for row_idx, row in enumerate(
                    _ws.iter_rows(min_row=15, max_row=_ws.max_row,
                                  values_only=True),
                    start=15,
                ):
                    for c in row:
                        if isinstance(c, str) and any(m in c for m in _cf_markers):
                            cashflow_sheets.append(_ws)
                            _cf_start_row = max(1, row_idx - 2)  # include header rows above
                            break
                    if cashflow_sheets:
                        break
                if cashflow_sheets:
                    break

        # Also try sheet-name-based matching as fallback for anything missed
        name_lower_map = {sn.lower(): wb[sn] for sn in wb.sheetnames}
        if not income_standalone and not income_annual and not income_unknown and not income_cumulative:
            for sn in wb.sheetnames:
                nl = sn.lower()
                if nl.startswith("pl") or "income" in nl:
                    income_unknown.append((sn, wb[sn]))
        if not bs_sheets:
            for sn in wb.sheetnames:
                nl = sn.lower()
                if nl.startswith("bs") or "balance" in nl:
                    bs_sheets.append(wb[sn])
        if not cashflow_sheets:
            for sn in wb.sheetnames:
                nl = sn.lower()
                if nl.startswith("cf") or "cash" in nl:
                    cashflow_sheets.append(wb[sn])

        # --- Phase 2: select best income statement sheet --------------------
        # Priority: annual (full-year) > standalone (3-month) > unknown
        # When BOTH annual and standalone exist in the same ZIP (e.g. BDMS annual
        # report that includes PL-T(3) for Q4 standalone AND PL-T(12) for full year),
        # use the annual sheet — its values are the canonical full-year figures.
        if income_annual:
            ws = income_annual[0][1]
            sections = self._parse_sheet_sections(ws)
            # Fallback: if annual sheet yields 0 data rows or only header/template rows
            # (e.g. BEM includes a blank "CI (1 statement)" template sheet classified
            # as annual), fall through to standalone instead.
            # A "real" annual sheet must have at least 3 rows with non-zero numeric values.
            _annual_rows = sections[0].get("rows", [])
            _real_data_rows = sum(
                1 for r in _annual_rows
                if r.get("label", "").strip()
                and r.get("label", "") not in ("หมายเหตุ", "Note", "Notes")
                and (r.get("consolidated_current") or 0) != 0
            )
            if _real_data_rows >= 3:
                result["income_statement"] = sections[0]
                if len(sections) > 1:
                    result["income_statement_cumulative"] = sections[1]
                # Store standalone Q4 if available
                if income_standalone:
                    result["income_statement_standalone"] = self._parse_sheet(
                        income_standalone[0][1])
            else:
                # Annual sheet has no data rows → fall through to standalone
                income_annual = []

        if not income_annual and income_standalone:
            # We have an explicitly standalone sheet — use it
            ws = income_standalone[0][1]
            sections = self._parse_sheet_sections(ws)
            result["income_statement"] = sections[0]
            if len(sections) > 1:
                result["income_statement_cumulative"] = sections[1]
        elif income_unknown:
            # Unknown period — may be a single combined sheet (KAMART style)
            # or annual report; use section splitter
            ws = income_unknown[0][1]
            sections = self._parse_sheet_sections(ws)
            result["income_statement"] = sections[0]
            if len(sections) > 1:
                result["income_statement_cumulative"] = sections[1]

        # If we found cumulative income sheet separately, store it
        if income_cumulative:
            result["income_statement_cumulative"] = self._parse_sheet(
                income_cumulative[0][1])

        # --- Phase 3: balance sheet (merge split sheets) --------------------
        if bs_sheets:
            merged = {"rows": [], "headers": [], "metadata": {}}
            for ws in bs_sheets:
                parsed = self._parse_sheet(ws)
                merged["rows"].extend(parsed.get("rows", []))
                if not merged["headers"] and parsed.get("headers"):
                    merged["headers"] = parsed["headers"]
                if not merged["metadata"].get("period_description"):
                    merged["metadata"] = parsed.get("metadata", {})
            result["balance_sheet"] = merged

        # --- Phase 4: cashflow (use first matching sheet) -------------------
        if cashflow_sheets:
            if _cf_start_row is not None:
                # Combined PL+CF sheet — parse only from the CF section onward
                result["cashflow"] = self._parse_sheet(
                    cashflow_sheets[0], min_row=_cf_start_row)
            else:
                result["cashflow"] = self._parse_sheet(cashflow_sheets[0])

        return result

    def _parse_sheet_sections(self, ws) -> list[dict]:
        """
        Parse an income statement sheet that may contain MULTIPLE sections.

        Quarterly XLSX files from SET often contain:
          Section 1 (rows 1-~45): Standalone quarter income statement
                    (e.g. "สำหรับงวดสามเดือน" = 3-month standalone)
          Section 2 (rows ~47-~73): Comprehensive income for standalone
          Section 3 (rows ~76-~118): Cumulative income statement
                    (e.g. "สำหรับงวดหกเดือน" = 6-month cumulative)
          Section 4 (rows ~119-~149): Comprehensive income for cumulative

        This method splits the sheet at section boundaries and returns
        a list of parsed sections.  Boundaries are detected by:
          - "งบกำไรขาดทุนเบ็ดเสร็จ" (comprehensive income header)
          - A second "งบกำไรขาดทุน" header after data rows
          - English equivalents: "Statement of comprehensive income"

        For annual XLSX (full-year) or Q1 XLSX (only one income statement
        section), this returns a single-element list.
        """
        # Read all rows as raw cell data
        all_rows_raw: list[list[tuple[str, object]]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            cells = []
            for c in row:
                try:
                    col_letter = c.column_letter
                except AttributeError:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(c.column)
                cells.append((col_letter, c.value))
            all_rows_raw.append(cells)

        # Detect section boundaries: find row indices where a NEW income
        # statement section begins (after "งบกำไรขาดทุนเบ็ดเสร็จ" or a
        # repeated "งบกำไรขาดทุน" header)
        _section_end_markers = (
            "งบกำไรขาดทุนเบ็ดเสร็จ",
            "Statement of comprehensive income",
            "Statements of comprehensive income",
        )
        _section_start_markers = (
            "งบกำไรขาดทุน",
            "Statement of profit or loss",
            "Statements of profit or loss",
            "Income statement",
            "Profit and loss",
        )

        # Find all section-start row indices
        section_starts: list[int] = []
        data_row_count = 0  # count rows with substantial data (label + number)

        for idx, row in enumerate(all_rows_raw):
            text_vals = [str(v).strip() for _, v in row if v and isinstance(v, str)]
            full_text = " ".join(text_vals)

            # Count how many numeric values this row has (to distinguish
            # data rows from header rows with stray numbers like page numbers)
            num_count = sum(1 for _, v in row if isinstance(v, (int, float)))
            has_label = any(len(t) > 5 for t in text_vals)
            if num_count >= 2 and has_label:
                data_row_count += 1

            # Only look for section boundaries after enough data rows
            # (at least 5 rows of actual data)
            if data_row_count < 5:
                continue

            # Check for section start markers
            for marker in _section_start_markers:
                if marker in full_text:
                    section_starts.append(idx)
                    break

            # Check for section end markers (comprehensive income = boundary)
            for marker in _section_end_markers:
                if marker in full_text:
                    if not section_starts:
                        section_starts.append(idx)
                    break

        if not section_starts:
            # Only one section — parse normally
            return [self._parse_sheet(ws)]

        # Split into sections and parse each
        # Section 1: rows 0 to first boundary
        # Section 2+: from boundary marker onwards
        # We find the NEXT income statement header after the boundary
        cut_row = section_starts[0]

        # Find the actual start of section 2 (the second "งบกำไรขาดทุน" header)
        section2_start = cut_row
        for idx in range(cut_row, len(all_rows_raw)):
            text_vals = [str(v).strip() for _, v in all_rows_raw[idx] if v and isinstance(v, str)]
            full_text = " ".join(text_vals)
            # Look for an income statement header (not comprehensive)
            is_income_header = False
            for marker in _section_start_markers:
                if marker in full_text:
                    is_income_header = True
                    break
            # Make sure it's not a comprehensive income header
            is_comprehensive = any(m in full_text for m in _section_end_markers)
            if is_income_header and not is_comprehensive and idx > cut_row:
                section2_start = idx
                break

        # Parse section 1 (standalone) using rows up to cut_row
        section1 = self._parse_sheet(ws, max_row=cut_row)

        # Parse section 2 (cumulative) using rows from section2_start
        section2 = self._parse_sheet(ws, min_row=section2_start + 1)

        return [section1, section2]

    def _parse_sheet(self, ws, min_row: int | None = None, max_row: int | None = None) -> dict:
        """
        Parse an Excel sheet into a dict of {label: {col_header: value}}.

        Auto-detects the column layout:
          - Label column: leftmost column containing Thai text data rows
          - Value columns: first pair of numeric columns after "งบการเงินรวม"
            (consolidated) header.  Falls back to any first two numeric cols.

        Supports both AOT-style (A=label, D/F=values) and
        PTT-style (B=label, G/I or H/J=values) layouts.

        Optional min_row/max_row (1-indexed) limit which rows are processed.
        """
        data = {"rows": [], "headers": [], "metadata": {}}

        effective_min = min_row if min_row is not None else 1
        effective_max = max_row if max_row is not None else ws.max_row

        # -- Read all rows ----------------------------------------------------
        all_rows: list[list[tuple[str, object]]] = []
        for row in ws.iter_rows(min_row=effective_min, max_row=effective_max, values_only=False):
            cells = []
            for c in row:
                try:
                    col_letter = c.column_letter
                except AttributeError:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(c.column)
                cells.append((col_letter, c.value))
            all_rows.append(cells)

        # -- Auto-detect label column and value columns -----------------------
        label_col = "A"
        val_col_cur = "D"
        val_col_prev = "F"
        sep_col_cur: str | None = None
        sep_col_prev: str | None = None

        # Detect "งบการเงินรวม" or "Consolidated" header row
        _date_keywords = ("ธันวาคม", "กันยายน", "มีนาคม", "มิถุนายน",
                          "December", "September", "March", "June")
        _conso_keywords = ("งบการเงินรวม", "ข้อมูลทางการเงินรวม",
                           "Consolidated", "consolidated")
        conso_header_col: str | None = None

        for row in all_rows:
            vals = {col: val for col, val in row if val and isinstance(val, str)}
            for col, val in vals.items():
                if any(kw in val for kw in _conso_keywords):
                    conso_header_col = col
                    break
            if conso_header_col:
                break

        # Find header row with dates (or Thai years like "2568" / 2568)
        import re as _re
        for row in all_rows:
            vals = {col: val for col, val in row if val}
            date_cols = []
            for col, val in vals.items():
                if isinstance(val, str) and any(dk in val for dk in _date_keywords):
                    date_cols.append(col)
                elif isinstance(val, (int, float)) and 2500 < val < 2700:
                    # Thai year as number (e.g. 2568)
                    date_cols.append(col)
                    vals[col] = str(int(val))
                elif isinstance(val, str):
                    # Thai year as string (e.g. "2568" or "พ.ศ. 2568")
                    stripped = val.strip()
                    # Extract 4-digit number from string
                    m = _re.search(r"(\d{4})", stripped)
                    if m:
                        y = int(m.group(1))
                        if 2500 < y < 2700:
                            date_cols.append(col)
                            vals[col] = str(y)
            if len(date_cols) >= 2:
                # Sort by column letter to pick consolidated (first pair)
                date_cols.sort()
                # If we know consolidated header column, pick the pair at/after it
                if conso_header_col and conso_header_col in date_cols:
                    idx = date_cols.index(conso_header_col)
                    val_col_cur = date_cols[idx]
                    val_col_prev = date_cols[idx + 1] if idx + 1 < len(date_cols) else date_cols[-1]
                else:
                    val_col_cur = date_cols[0]
                    val_col_prev = date_cols[1]
                data["headers"] = [str(vals.get(val_col_cur, "")), str(vals.get(val_col_prev, ""))]
                # Separate entity columns (next pair after consolidated)
                remaining = [c for c in date_cols[2:] if c > val_col_prev]
                if len(remaining) >= 2:
                    sep_col_cur = remaining[0]
                    sep_col_prev = remaining[1]
                    data["headers"].extend([str(vals.get(sep_col_cur, "")), str(vals.get(sep_col_prev, ""))])
                break

        # --- Verify val_col_cur actually contains numeric data ---------
        # Some filings use merged cells in the header row so the date text
        # lands in col C but the actual numbers are in col D (one col right).
        # Check if the detected val_col_cur has any numeric rows; if not,
        # shift both val columns right by 1 up to 3 times.
        from openpyxl.utils import column_index_from_string, get_column_letter
        def _has_numeric_data(rows, col):
            for r in rows[5:]:
                v = dict(r).get(col)
                if isinstance(v, (int, float)):
                    return True
            return False

        for _shift in range(1, 4):
            if _has_numeric_data(all_rows, val_col_cur):
                break
            # Shift both columns right by 1
            try:
                idx_cur = column_index_from_string(val_col_cur)
                idx_prev = column_index_from_string(val_col_prev)
                val_col_cur = get_column_letter(idx_cur + 1)
                val_col_prev = get_column_letter(idx_prev + 1)
                if sep_col_cur:
                    idx_sc = column_index_from_string(sep_col_cur)
                    idx_sp = column_index_from_string(sep_col_prev)
                    sep_col_cur = get_column_letter(idx_sc + 1)
                    sep_col_prev = get_column_letter(idx_sp + 1)
            except Exception:
                break

        # Detect label columns: find columns with Thai (or English) text strings.
        # Some filings put main labels in col B and subtotals in col C.
        _thai_range = range(0x0E00, 0x0E80)
        def _is_thai(s: str) -> bool:
            return any(ord(ch) in _thai_range for ch in s[:20])

        col_thai_count: dict[str, int] = {}
        col_text_count: dict[str, int] = {}  # fallback: any text (for English XLSX)
        for row in all_rows:
            for col, val in row:
                if isinstance(val, str) and val.strip():
                    if _is_thai(val):
                        col_thai_count[col] = col_thai_count.get(col, 0) + 1
                    elif len(val.strip()) > 3 and val.strip()[0].isalpha():
                        col_text_count[col] = col_text_count.get(col, 0) + 1

        # Prefer Thai text columns; fall back to English text columns
        label_count = col_thai_count if col_thai_count else col_text_count
        if label_count:
            label_col = max(label_count, key=label_count.get)  # type: ignore

        # Build list of all possible label columns (sorted by text count)
        label_cols = sorted(
            [c for c, n in label_count.items() if c < val_col_cur],
            key=lambda c: label_count.get(c, 0), reverse=True
        )
        if not label_cols:
            label_cols = [label_col]

        # If label col and val_col are the same, recalculate
        if label_col == val_col_cur:
            for row in all_rows[5:15]:
                row_dict = {col: val for col, val in row}
                for col in sorted(row_dict.keys()):
                    if col > label_col and isinstance(row_dict[col], (int, float)):
                        val_col_cur = col
                        break
                if val_col_cur != label_col:
                    break

        # -- Capture period description and unit --------------------------------
        for row in all_rows:
            row_dict = {col: val for col, val in row}
            for check_col in label_cols + ["A", "B"]:
                txt = row_dict.get(check_col, "")
                if isinstance(txt, str):
                    # Thai period descriptions
                    if "สำหรับ" in txt and "สิ้นสุด" in txt:
                        data["metadata"]["period_description"] = txt.strip()
                        break
                    if "ณ วัน" in txt:
                        data["metadata"]["period_description"] = txt.strip()
                        break
                    # English period descriptions
                    if "For the" in txt and ("period ended" in txt or "year ended" in txt):
                        data["metadata"]["period_description"] = txt.strip()
                        break
                    if "As at " in txt:
                        data["metadata"]["period_description"] = txt.strip()
                        break
            if data["metadata"].get("period_description"):
                break

        # Detect unit (พันบาท = thousands, ล้านบาท = millions, บาท = baht)
        for row in all_rows[:15]:
            for col, val in row:
                if isinstance(val, str):
                    txt = val.strip().lower()
                    if "พันบาท" in txt or "thousand" in txt:
                        data["metadata"]["unit"] = "thousands"
                        break
                    if "ล้านบาท" in txt or "million" in txt:
                        data["metadata"]["unit"] = "millions"
                        break
                    if "(หน่วย: บาท)" in val or "(Unit: Baht)" in val:
                        data["metadata"]["unit"] = "baht"
                        break
            if data["metadata"].get("unit"):
                break

        # -- Parse data rows ---------------------------------------------------
        _skip_keywords = (
            # Thai
            "งบ", "ณ วัน", "หน่วย", "หมายเหตุประกอบ",
            "(ยัง", "(ตรวจสอบ", "แต่สอบทาน", "งบการเงิน",
            # English
            "Company", "Statement", "As at", "Unit:", "(Unit",
            "(Unaudited", "(Audited", "Consolidated", "Separate",
            "Note", "For the",
        )
        # These prefixes indicate financial data rows even if they contain "บริษัท"
        _data_prefixes = (
            "ส่วนที่เป็นของ", "การแบ่งปัน",
        )
        for row in all_rows:
            row_dict = {col: val for col, val in row}

            # Find label: check all label columns (B, C, A, etc.)
            label = None
            for lc in label_cols:
                v = row_dict.get(lc)
                if v and isinstance(v, str) and v.strip():
                    label = v.strip()
                    break
            cur_val = row_dict.get(val_col_cur)
            prev_val = row_dict.get(val_col_prev)

            if not label:
                continue

            # Skip header/metadata rows
            # "บริษัท" alone is a header indicator, but only when the label STARTS with it
            # (not when it appears mid-label, e.g. "ส่วนที่เป็นของผู้ถือหุ้นของบริษัทฯ")
            is_data_row = any(label.startswith(p) for p in _data_prefixes)
            if is_data_row:
                is_header = False
            else:
                is_header = any(sk in label for sk in _skip_keywords)
                if not is_header and label.startswith("บริษัท"):
                    is_header = True
                if not is_header and "สำหรับ" in label and label.startswith("สำหรับ"):
                    is_header = True
            if is_header:
                if "บริษัท" in label:
                    data["metadata"]["company"] = label
                continue

            def _to_num(v):
                """Convert cell value to float, handling string numbers and dashes."""
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                if isinstance(v, str):
                    s = v.strip().replace(",", "").replace("\xa0", "").replace(" ", "")
                    if s in ("-", "–", "—", ""):
                        return 0.0
                    try:
                        return float(s)
                    except ValueError:
                        return None
                return None

            cur_num = _to_num(cur_val)
            if cur_num is not None:
                prev_num = _to_num(prev_val)
                entry = {"label": label, "consolidated_current": cur_num, "consolidated_prev": prev_num}
                if sep_col_cur:
                    val_h = _to_num(row_dict.get(sep_col_cur))
                    val_j = _to_num(row_dict.get(sep_col_prev))
                    if val_h is not None:
                        entry["separate_current"] = val_h
                        entry["separate_prev"] = val_j
                data["rows"].append(entry)

        return data

    # ------------------------------------------------------------------
    # XLSX → quarterly summary extraction
    # ------------------------------------------------------------------
    # Label keywords → field mapping for key-item extraction.
    # Each tuple: (search_keywords, field_name, match_mode)
    #   match_mode: "startswith" or "contains"
    # Each tuple: (keywords_list, field_name, exact_match)
    #   exact_match=True → label must equal one of the keywords exactly
    #   exact_match=False → label.startswith(keyword) is enough
    _INCOME_KEYS: list[tuple[list[str], str, bool]] = [
        # Thai + English label variants
        # "sales" = revenue from core business (selling goods/services)
        (["รวมรายได้จากการขาย", "รายได้จากการขายสินค้าและบริการ",
          "รายได้จากการขายและการให้บริการ",
          "รายได้จากการขายสินค้าและการให้บริการ",
          "รายได้จากการขายและบริการ",
          "รายได้จากการให้บริการ",       # BEM (service revenue)
          "รายได้จากการขาย",             # GFPT (sales revenue)
          "Total revenues from sales or services", "Total revenues from sales",
          "Revenue from sales and services", "Revenue from sale",
          "Revenue from rendering of services"], "sales", False),
        (["รวมรายได้อื่น", "รายได้อื่น",
          "Total other income", "Other income"], "other_revenue", False),
        (["รวมรายได้",
          "Total revenues", "Total revenue"], "total_revenue", False),
        (["รวมค่าใช้จ่าย",
          "Total expenses", "Total cost and expenses"], "total_expense", False),
        (["กำไรขั้นต้น", "กำไร(ขาดทุน)ขั้นต้น",
          "กำไร (ขาดทุน) ขั้นต้น",
          "Gross profit", "Gross profit (loss)"], "gross_profit", False),
        (["กำไรจากการดำเนินงาน", "กำไร(ขาดทุน)จากการดำเนินงาน",
          "กำไร (ขาดทุน) จากการดำเนินงาน",  # MINT style with spaces
          "กำไรจากกิจกรรมดำเนินงาน", "กำไร(ขาดทุน)จากกิจกรรมดำเนินงาน",
          "Profit from operating activities", "Operating profit",
          "Operating profit (loss)"], "operating_profit", False),
        (["กำไรก่อนภาษี", "กำไร(ขาดทุน)ก่อนภาษี",
          "กำไรก่อนค่าใช้จ่ายภาษีเงินได้", "กำไร(ขาดทุน)ก่อนค่าใช้จ่ายภาษีเงินได้",
          "กำไร (ขาดทุน) ก่อนภาษี",               # MINT style
          "กำไร (ขาดทุน) ก่อนค่าใช้จ่ายภาษีเงินได้",  # MINT style
          "ขาดทุนก่อนภาษี",                         # Loss-making quarters (AOT COVID era)
          "Profit before income tax", "Profit (loss) before income tax",
          "Loss before income tax",
          "Profit before tax"], "profit_before_tax", False),
        (["กำไรสำหรับงวด", "กำไรสำหรับปี", "กำไร(ขาดทุน)สำหรับงวด",
          "กำไร(ขาดทุน)สำหรับปี", "ขาดทุนสำหรับงวด", "ขาดทุนสำหรับปี",
          "กำไร(ขาดทุน)รวมสำหรับงวด", "กำไร(ขาดทุน)รวมสำหรับปี",  # INET style
          "กำไร (ขาดทุน) สำหรับ",   # MINT style with spaces
          "กำไรสุทธิสำหรับปี", "กำไรสุทธิสำหรับงวด",              # KTC style
          "ขาดทุนสุทธิสำหรับปี", "ขาดทุนสุทธิสำหรับงวด",
          "กำไรสุทธิ",                                             # KBANK style (exact)
          "Profit for the period", "Profit for the year",
          "Profit (loss) for the period", "Profit (loss) for the year",
          "Net profit for the period", "Net profit for the year",
          "Loss for the period", "Loss for the year"], "net_profit", False),
        (["ส่วนที่เป็นของผู้ถือหุ้นของบริษัท",   # covers บริษัทฯ, บริษัทใหญ่, etc. (startswith)
          "ส่วนที่เป็นของบริษัท",                # TU style (no ผู้ถือหุ้น prefix)
          "ส่วนที่เป็นของผู้เป็นเจ้าของ",          # MINT style
          "Equity holders of the Company", "Owners of the parent"], "ni_owners", False),
        (["กำไรต่อหุ้นขั้นพื้นฐาน", "ขาดทุนต่อหุ้นขั้นพื้นฐาน",
          "Basic earnings per share", "Basic earnings (loss) per share",
          "Earnings per share"], "eps", False),
        (["ต้นทุนทางการเงิน",
          "Finance cost", "Finance costs"], "finance_cost", False),
        (["ค่าใช้จ่ายภาษีเงินได้",
          "รายได้ (ค่าใช้จ่าย) ภาษีเงินได้",
          "รายได้(ค่าใช้จ่าย)ภาษีเงินได้",
          "(ค่าใช้จ่าย)รายได้ภาษีเงินได้",
          "(ค่าใช้จ่าย) รายได้ภาษีเงินได้",
          "ค่าใช้จ่าย (รายได้) ภาษีเงินได้",
          "ค่าใช้จ่าย(รายได้)ภาษีเงินได้",       # INET annual style (no spaces)
          "รายได้ภาษีเงินได้",                   # Tax benefit (loss-making quarters)
          "ภาษีเงินได้",
          "Income tax expense", "Income tax",
          "Tax expense", "Tax benefit"], "tax_expense", False),
        (["ค่าเสื่อมราคาและค่าตัดจำหน่าย",
          "Depreciation and amortisation", "Depreciation and amortization"], "depreciation", False),
    ]
    _BALANCE_KEYS: list[tuple[list[str], str, bool]] = [
        (["เงินสดและรายการเทียบเท่าเงินสด", "เงินสดและเงินฝากธนาคาร",
          "Cash and cash equivalents", "Cash and bank balances",
          "เงินสด"], "cash", False),
        (["สินค้าคงเหลือ", "สินค้าคงคลัง",
          "Inventories", "Inventory"], "inventories", False),
        (["เงินลงทุนระยะสั้น", "เงินลงทุนชั่วคราว",
          "สินทรัพย์ทางการเงินหมุนเวียนอื่น",
          "Short-term investments", "Current investments",
          "Other current financial assets"], "short_term_investments", False),
        (["ลูกหนี้การค้าและลูกหนี้อื่น", "ลูกหนี้การค้า",
          "ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น",
          "Trade and other receivables", "Trade receivables",
          "Trade and other current receivables",
          "Trade accounts receivable"], "trade_receivables", False),
        (["รวมสินทรัพย์หมุนเวียน",
          "Total current assets"], "current_assets", False),
        (["รวมสินทรัพย์ไม่หมุนเวียน",
          "Total non-current assets"], "non_current_assets", False),
        (["รวมสินทรัพย์",
          "Total assets"], "total_assets", True),
        (["รวมหนี้สินหมุนเวียน",
          "Total current liabilities"], "current_liabilities", False),
        (["รวมหนี้สินไม่หมุนเวียน",
          "Total non-current liabilities"], "non_current_liabilities", False),
        (["รวมหนี้สิน",
          "Total liabilities"], "total_liabilities", True),
        (["รวมส่วนของผู้ถือหุ้น", "ส่วนของผู้ถือหุ้นของบริษัท",
          "รวมส่วนของเจ้าของ",        # GPSC/MINT/WHA style
          "Total shareholders' equity", "Total equity",
          "Equity attributable to owners of the Company"], "equity", False),
    ]
    _CASHFLOW_KEYS: list[tuple[list[str], str, bool]] = [
        (["เงินสดสุทธิได้มาจากกิจกรรมดำเนินงาน", "เงินสดสุทธิใช้ไปในกิจกรรมดำเนินงาน",
          "เงินสดสุทธิจากกิจกรรมดำเนินงาน",
          "เงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมดำเนินงาน",
          "เงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมดำเนินงาน",
          "เงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมดำเนินงาน",
          "เงินสดสุทธิได้มาจากการดำเนินงาน",
          "เงินสดสุทธิใช้ไปในการดำเนินงาน",
          "เงินสดสุทธิจากการดำเนินงาน",
          # INET style: "(ใช้ไป)" (no ใน) / "เงินสุทธิ" (no สด)
          "เงินสดสุทธิได้มา(ใช้ไป)จากกิจกรรมดำเนินงาน",
          "เงินสดสุทธิได้มา (ใช้ไป) จากกิจกรรมดำเนินงาน",
          "เงินสุทธิได้มา(ใช้ไป)จากกิจกรรมดำเนินงาน",
          "เงินสุทธิได้มา (ใช้ไป) จากกิจกรรมดำเนินงาน",
          "เงินสุทธิได้มาจากกิจกรรมดำเนินงาน",
          "เงินสุทธิใช้ไปในกิจกรรมดำเนินงาน",
          # CPALL style: "กระแสเงินสด" prefix
          "กระแสเงินสดสุทธิได้มาจากกิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิใช้ไปในกิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิจากกิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมดำเนินงาน",
          "กระแสเงินสดสุทธิได้มาจากการดำเนินงาน",
          "กระแสเงินสดสุทธิใช้ไปในการดำเนินงาน",
          "กระแสเงินสดสุทธิจากการดำเนินงาน",
          "Net cash provided by operating activities",
          "Net cash generated from operating activities",
          "Net cash from operating activities",
          "Net cash flows from operating activities",
          "Net cash flows from (used in) operating activities",
          "Net cash flows used in operating activities",
          "Net cash provided by (used in) operating activities"], "cf_operating", False),
        (["เงินสดสุทธิได้มาจากกิจกรรมลงทุน", "เงินสดสุทธิใช้ไปในกิจกรรมลงทุน",
          "เงินสดสุทธิจากกิจกรรมลงทุน",
          "เงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมลงทุน",
          "เงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมลงทุน",
          "เงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมลงทุน",
          "เงินสดสุทธิได้มาจากการลงทุน",
          "เงินสดสุทธิใช้ไปในการลงทุน",
          "เงินสดสุทธิจากการลงทุน",
          # INET style: "เงินสุทธิ" (no สด) + "(ใช้ไป)" (no ใน)
          "เงินสุทธิได้มา(ใช้ไป)จากกิจกรรมลงทุน",
          "เงินสุทธิได้มา (ใช้ไป) จากกิจกรรมลงทุน",
          "เงินสดสุทธิได้มา(ใช้ไป)จากกิจกรรมลงทุน",
          "เงินสดสุทธิได้มา (ใช้ไป) จากกิจกรรมลงทุน",
          # CPALL style
          "กระแสเงินสดสุทธิได้มาจากกิจกรรมลงทุน",
          "กระแสเงินสดสุทธิใช้ไปในกิจกรรมลงทุน",
          "กระแสเงินสดสุทธิจากกิจกรรมลงทุน",
          "กระแสเงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมลงทุน",
          "กระแสเงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมลงทุน",
          "กระแสเงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมลงทุน",
          "กระแสเงินสดสุทธิได้มาจากการลงทุน",
          "กระแสเงินสดสุทธิใช้ไปในการลงทุน",
          "กระแสเงินสดสุทธิจากการลงทุน",
          "Net cash used in investing activities",
          "Net cash from investing activities",
          "Net cash flows from investing activities",
          "Net cash flows from (used in) investing activities",
          "Net cash flows used in investing activities",
          "Net cash provided by (used in) investing activities"], "cf_investing", False),
        (["เงินสดสุทธิได้มาจากกิจกรรมจัดหาเงิน", "เงินสดสุทธิใช้ไปในกิจกรรมจัดหาเงิน",
          "เงินสดสุทธิจากกิจกรรมจัดหาเงิน",
          "เงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมจัดหาเงิน",
          "เงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมจัดหาเงิน",
          "เงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมจัดหาเงิน",
          # INET style: "(ใช้ไป)" (no ใน)
          "เงินสดสุทธิได้มา(ใช้ไป)จากกิจกรรมจัดหาเงิน",
          "เงินสดสุทธิได้มา (ใช้ไป) จากกิจกรรมจัดหาเงิน",
          "เงินสุทธิได้มา(ใช้ไป)จากกิจกรรมจัดหาเงิน",
          "เงินสุทธิได้มา (ใช้ไป) จากกิจกรรมจัดหาเงิน",
          # CPALL style
          "กระแสเงินสดสุทธิได้มาจากกิจกรรมจัดหาเงิน",
          "กระแสเงินสดสุทธิใช้ไปในกิจกรรมจัดหาเงิน",
          "กระแสเงินสดสุทธิจากกิจกรรมจัดหาเงิน",
          "กระแสเงินสดสุทธิได้มาจาก(ใช้ไปใน)กิจกรรมจัดหาเงิน",
          "กระแสเงินสดสุทธิได้มาจาก (ใช้ไปใน) กิจกรรมจัดหาเงิน",
          "กระแสเงินสดสุทธิจาก (ใช้ไปใน) กิจกรรมจัดหาเงิน",
          "Net cash used in financing activities",
          "Net cash from financing activities",
          "Net cash flows from financing activities",
          "Net cash flows from (used in) financing activities",
          "Net cash flows used in financing activities",
          "Net cash provided by (used in) financing activities"], "cf_financing", False),
        (["เงินสดและรายการเทียบเท่าเงินสดเพิ่มขึ้น", "เงินสดและรายการเทียบเท่าเงินสดลดลง",
          "เงินสดและรายการเทียบเท่าเงินสดเพิ่มขึ้น(ลดลง)",
          "เงินสดและรายการเทียบเท่าเงินสดเพิ่มขึ้น (ลดลง)",
          # Reordered: "(ลดลง) เพิ่มขึ้น" (CPALL style)
          "เงินสดและรายการเทียบเท่าเงินสด (ลดลง)",
          "เงินสดและรายการเทียบเท่าเงินสด(ลดลง)",
          # With สุทธิ before เพิ่มขึ้น (KAMART style)
          "เงินสดและรายการเทียบเท่าเงินสดสุทธิเพิ่มขึ้น",
          "เงินสดและรายการเทียบเท่าเงินสดสุทธิลดลง",
          "เพิ่มขึ้น (ลดลง) สุทธิ",
          "Net increase in cash and cash equivalents",
          "Net decrease in cash and cash equivalents",
          "Net increase (decrease) in cash"], "cf_net", False),
    ]

    # -----------------------------------------------------------------------
    # Special items: non-cash / non-operating items that are REVERSED OUT
    # in the CF operating section (indirect method).
    # These represent items between operating_profit and profit_before_tax
    # in the IS — i.e. the true "extraordinary / non-recurring" items.
    #
    # Convention: CF indirect method เริ่มจาก PBT แล้ว บวกกลับ/หัก non-cash
    # รายการที่ "หัก" (เป็นลบ) ใน adjustment section = รายการที่ไม่ใช่เงินสด
    # ที่รับรู้เป็นกำไรใน IS → เราต้องนำไปหักออกจาก core profit
    #
    # เก็บเป็น list ของ field names ที่ต้องการ parse จาก CF rows
    # (แต่ละ field จะรวมค่าไว้ใน special_items dict)
    # -----------------------------------------------------------------------
    _CF_SPECIAL_ITEMS_KEYS: list[tuple[list[str], str, bool]] = [
        # กำไร/ขาดทุนจากการขายสินทรัพย์
        (["กำไรจากการจำหน่ายสินทรัพย์", "กำไรจากการขายสินทรัพย์",
          "กำไรจากการจำหน่ายที่ดิน", "กำไรจากการขายที่ดิน",
          "Gain on disposal of assets", "Gain on sale of assets",
          "Gain on disposal of property"], "gain_on_disposal_assets", False),
        # กำไร/ขาดทุนจากเงินลงทุน (ราคาตลาด/fair value)
        (["กำไรจากการเปลี่ยนแปลงในมูลค่ายุติธรรมของ",
          "กำไรจากการเปลี่ยนแปลงมูลค่ายุติธรรมของเงินลงทุน",
          "กำไรจากการเปลี่ยนแปลงมูลค่าเงินลงทุน",
          "Gain from change in fair value",
          "Gain on change in fair value of investments"], "gain_on_fv_investments", False),
        # กำไร/ขาดทุนจากอสังหาริมทรัพย์เพื่อการลงทุน
        (["กำไรจากการเปลี่ยนแปลงในมูลค่ายุติธรรมของ\n   อสังหาริมทรัพย์",
          "อสังหาริมทรัพย์เพื่อการลงทุน",
          "Gain from change in fair value of investment property",
          "Gain on investment property"], "gain_on_fv_inv_property", False),
        # กำไรจากการขายเงินลงทุน
        (["กำไรจากการขายเงินลงทุน", "กำไรจากการจำหน่ายเงินลงทุน",
          "Gain on sale of investments", "Gain on disposal of investments",
          "Gain on disposal of subsidiaries"], "gain_on_sale_investments", False),
        # ส่วนแบ่งกำไรจากเงินลงทุนในบริษัทร่วม/ร่วมค้า
        (["ส่วนแบ่งกำไรจากเงินลงทุนในการร่วมค้า",
          "ส่วนแบ่งกำไรจากเงินลงทุนในบริษัทร่วม",
          "ส่วนแบ่งกำไร(ขาดทุน)จากเงินลงทุนในการร่วมค้า",
          "Share of profit from investments in associates",
          "Share of profit (loss) of associates",
          "Share of profits of associates and joint ventures"], "share_of_profit_associates", False),
        # ดอกเบี้ยรับ / รายได้ทางการเงิน
        (["ดอกเบี้ยรับ", "รายได้ดอกเบี้ย",
          "Interest received", "Interest income",
          "Finance income received"], "interest_income_cf", False),
        # เงินปันผลรับ
        (["เงินปันผลรับ", "Dividend received", "Dividends received"], "dividend_received", False),
        # กำไรจากอัตราแลกเปลี่ยนที่ยังไม่เกิดขึ้น (unrealised FX)
        (["ขาดทุนจากอัตราแลกเปลี่ยนที่ยังไม่เกิดขึ้น",
          "กำไรจากอัตราแลกเปลี่ยนที่ยังไม่เกิดขึ้น",
          "Unrealized loss on exchange", "Unrealized gain on exchange",
          "Unrealised foreign exchange"], "unrealised_fx", False),
        # กำไรจากตราสารอนุพันธ์
        (["กำไรจากตราสารอนุพันธ์", "Gain on derivatives"], "gain_on_derivatives", False),
    ]

    # Fields that must appear AFTER profit_before_tax in the income statement.
    # For these fields the FIRST match after PBT wins (not the largest-abs-value),
    # because some sheets have a mis-labeled or structurally duplicated row earlier
    # in the statement that happens to share the same keyword (e.g. CPALL Q9/2024
    # has two rows matching "ค่าใช้จ่ายภาษีเงินได้" — the first is near EBIT with a
    # much larger value; the real tax expense appears just before net_profit).
    _POST_PBT_FIELDS: frozenset[str] = frozenset({"tax_expense"})

    # Keywords that mark the profit_before_tax anchor row (subset of _INCOME_KEYS).
    _PBT_ANCHOR_KWS: tuple[str, ...] = (
        "กำไรก่อนภาษี", "กำไรก่อนค่าใช้จ่ายภาษีเงินได้",
        "กำไร(ขาดทุน)ก่อนภาษี", "กำไร(ขาดทุน)ก่อนค่าใช้จ่ายภาษีเงินได้",
        "กำไร (ขาดทุน) ก่อนภาษี", "กำไร (ขาดทุน) ก่อนค่าใช้จ่ายภาษีเงินได้",
        "ขาดทุนก่อนภาษี",                                   # Loss-making quarters
        "Profit before income tax", "Profit (loss) before income tax",
        "Loss before income tax",
        "Profit before tax",
    )

    def _extract_key_items(self, rows: list[dict], key_map: list) -> dict:
        """
        From a list of parsed sheet rows, extract key summary items.
        Returns {field_name: {"current": val, "prev": val}}.

        Uses LAST match for each field so that grand-totals (which appear
        later in the sheet) override sub-totals.

        Non-zero preference: if the field already has a non-zero value and
        the new match produces zero, keep the existing non-zero value.
        This prevents zero-value duplicate rows (e.g. "รายได้จากการขาย" = 0
        in a service company) from overriding the real non-zero revenue row.

        Post-PBT fields (e.g. tax_expense): once profit_before_tax has been
        seen, the FIRST non-zero match wins and is locked — subsequent duplicate
        label rows (structurally earlier in the sheet would have been picked up
        before PBT anyway) cannot override it.
        """
        result = {}
        pbt_seen = False          # have we passed the profit_before_tax row?
        post_pbt_locked: set[str] = set()  # fields already locked after PBT

        for entry in rows:
            label: str = entry.get("label", "")

            # Track whether we have passed the profit_before_tax anchor.
            if not pbt_seen:
                for kw in self._PBT_ANCHOR_KWS:
                    if label.startswith(kw):
                        pbt_seen = True
                        break

            for keywords, field, exact in key_map:
                for kw in keywords:
                    matched = (label == kw) if exact else label.startswith(kw)
                    if matched:
                        new_curr = entry.get("consolidated_current", 0) or 0
                        new_prev = entry.get("consolidated_prev", 0) or 0
                        existing = result.get(field)

                        # Post-PBT fields: lock on the first non-zero match
                        # that appears after the profit_before_tax anchor row.
                        if field in self._POST_PBT_FIELDS:
                            if field in post_pbt_locked:
                                break  # already locked — ignore later duplicates
                            if pbt_seen and (new_curr != 0 or new_prev != 0):
                                result[field] = {"current": new_curr, "prev": new_prev}
                                post_pbt_locked.add(field)
                                break
                            # Before PBT: store tentatively (may be overridden
                            # once we cross PBT), using larger-abs-value rule.
                            if existing:
                                keep_curr = new_curr if abs(new_curr) > abs(existing["current"]) else existing["current"]
                                keep_prev = new_prev if abs(new_prev) > abs(existing["prev"]) else existing["prev"]
                                result[field] = {"current": keep_curr, "prev": keep_prev}
                            else:
                                result[field] = {"current": new_curr, "prev": new_prev}
                            break

                        if existing:
                            # For each column: prefer the larger absolute value.
                            # This handles cases like INET where "รายได้จากการให้บริการ"
                            # (large) and "รายได้จากการขาย" (small/zero) both match
                            # "sales" — we want the larger one to win regardless of order.
                            keep_curr = new_curr if abs(new_curr) > abs(existing["current"]) else existing["current"]
                            keep_prev = new_prev if abs(new_prev) > abs(existing["prev"]) else existing["prev"]
                            result[field] = {"current": keep_curr, "prev": keep_prev}
                        else:
                            result[field] = {"current": new_curr, "prev": new_prev}
                        break
        return result

    def _extract_special_items_from_cf(self, cf_rows: list[dict]) -> dict:
        """
        Extract non-operating / extraordinary items from the CF indirect method
        adjustment section.

        In the indirect method the CF statement starts from profit-before-tax
        and reverses out non-cash / non-operating items.  Items that are
        REVERSED OUT (shown as negative) are gains that were included in PBT
        but are not part of core operating income — e.g. fair value gains,
        share of profit from associates, interest income.

        IMPORTANT: Only search within the operating activities adjustment section
        (rows before the "net cash from operating" subtotal).  This prevents
        picking up cash receipts in investing/financing sections which have the
        same labels (e.g. "ดอกเบี้ยรับ" appears both as an adjustment and as
        an actual cash collected item later).

        Returns dict: {field_name: {"current": val, "prev": val}}
        where values carry the RAW CF sign (negative = gain reversed out,
        positive = loss reversed in).  Callers should use the raw sign
        directly: a negative value means it was a gain in the IS.
        """
        # Find the end of the operating section: the row that contains
        # "net cash from/used in operating activities"
        _operating_total_keywords = (
            "เงินสดสุทธิได้มาจากกิจกรรมดำเนินงาน",
            "เงินสดสุทธิใช้ไปในกิจกรรมดำเนินงาน",
            "เงินสดสุทธิจากกิจกรรมดำเนินงาน",
            "กระแสเงินสดสุทธิได้มาจากกิจกรรมดำเนินงาน",
            "กระแสเงินสดสุทธิใช้ไปในกิจกรรมดำเนินงาน",
            "Net cash provided by operating activities",
            "Net cash generated from operating activities",
            "Net cash flows from operating activities",
            "Net cash flows used in operating activities",
        )
        operating_end_idx = len(cf_rows)
        for i, row in enumerate(cf_rows):
            label = row.get("label", "")
            if any(kw in label for kw in _operating_total_keywords):
                operating_end_idx = i + 1  # include this row
                break

        # Only search in the operating section
        adj_rows = cf_rows[:operating_end_idx]
        return self._extract_key_items(adj_rows, self._CF_SPECIAL_ITEMS_KEYS)

    def extract_quarterly_summary(self, xlsx_bytes: bytes) -> dict:
        """
        Parse XLSX and return a structured quarterly summary dict.

        Returns::

            {
                "period_description": "สำหรับงวดสามเดือน...",
                "headers": ["31 ธันวาคม 2568", "31 ธันวาคม 2567"],
                "income": {field: {"current": ..., "prev": ...}, ...},
                "balance": {field: {"current": ..., "prev": ...}, ...},
                "cashflow": {field: {"current": ..., "prev": ...}, ...},
                "special_items_from_cf": {field: {"current": ..., "prev": ...}, ...},
                "cashflow_cumulative": True/False,
            }

        Income statement values are STANDALONE quarterly figures (the parser
        now correctly splits multi-section sheets and uses only the first
        standalone section).

        Cashflow values may be CUMULATIVE for Q2/Q3 filings (cashflow sheets
        only have cumulative data).  The "cashflow_cumulative" flag is set
        to True when the cashflow period is > 3 months.

        special_items_from_cf contains non-operating items reversed out in the
        CF indirect method adjustment section (raw CF sign: negative = IS gain).
        """
        parsed = self.parse_financial_xlsx(xlsx_bytes)
        # income_statement now contains only the standalone section
        inc_rows = parsed.get("income_statement", {}).get("rows", [])
        bal_rows = parsed.get("balance_sheet", {}).get("rows", [])
        cf_rows = parsed.get("cashflow", {}).get("rows", [])

        headers = (
            parsed.get("income_statement", {}).get("headers")
            or parsed.get("balance_sheet", {}).get("headers")
            or []
        )
        period_desc = (
            parsed.get("income_statement", {}).get("metadata", {}).get("period_description", "")
            or parsed.get("balance_sheet", {}).get("metadata", {}).get("period_description", "")
        )

        income = self._extract_key_items(inc_rows, self._INCOME_KEYS)
        balance = self._extract_key_items(bal_rows, self._BALANCE_KEYS)
        cashflow = self._extract_key_items(cf_rows, self._CASHFLOW_KEYS)
        special_items = self._extract_special_items_from_cf(cf_rows)

        # Fallback 1: compute total_revenue from sales + other_revenue if missing
        if "total_revenue" not in income and "sales" in income:
            s = income["sales"]
            o = income.get("other_revenue", {"current": 0, "prev": 0})
            income["total_revenue"] = {
                "current": (s.get("current", 0) or 0) + (o.get("current", 0) or 0),
                "prev": (s.get("prev", 0) or 0) + (o.get("prev", 0) or 0),
            }

        # Fallback 2: ถ้า total_revenue น้อยกว่า sum ของ positive "รายได้" rows อย่างมีนัยสำคัญ
        # (เช่น WHA มีหลาย revenue stream ไม่มี "รวมรายได้" line)
        # → sum ทุก row ที่ label ขึ้นต้นด้วย "รายได้" และ value > 0 เป็น total_revenue
        cur_tr = (income.get("total_revenue", {}) or {}).get("current", 0) or 0
        prev_tr = (income.get("total_revenue", {}) or {}).get("prev", 0) or 0
        rev_prefix = ("รายได้", "Revenue from", "Total revenues from")
        sum_rev_cur = sum(
            float(r.get("consolidated_current", 0) or 0)
            for r in inc_rows
            if any(str(r.get("label", "")).startswith(p) for p in rev_prefix)
            and float(r.get("consolidated_current", 0) or 0) > 0
        )
        sum_rev_prev = sum(
            float(r.get("consolidated_prev", 0) or 0)
            for r in inc_rows
            if any(str(r.get("label", "")).startswith(p) for p in rev_prefix)
            and float(r.get("consolidated_prev", 0) or 0) > 0
        )
        if sum_rev_cur > cur_tr * 1.05 and sum_rev_cur > 0:
            income["total_revenue"] = {"current": sum_rev_cur, "prev": sum_rev_prev}

        # Detect unit from any parsed sheet's metadata
        unit = "baht"  # default assumption
        for key in ("income_statement", "balance_sheet", "cashflow"):
            sheet_meta = parsed.get(key, {}).get("metadata", {})
            if sheet_meta.get("unit"):
                unit = sheet_meta["unit"]
                break

        # Detect if cashflow is cumulative (Q2/Q3 filings have cumulative
        # cashflow sheets — e.g. "สำหรับงวดหกเดือน" or "สำหรับงวดเก้าเดือน")
        cf_period_desc = parsed.get("cashflow", {}).get("metadata", {}).get("period_description", "")
        cf_p = _parse_thai_period_static(cf_period_desc)
        cf_months = cf_p.get("months_count", 0)
        cashflow_cumulative = cf_months > 3

        # Extract 9M cumulative income (present in Q3 ZIPs as a separate sheet)
        # Used by _add_q4_rows_from_xlsx when this is a q4_helper entry:
        #   Q4 = FY(12M) − income_9m(9M)
        cum_rows = parsed.get("income_statement_cumulative", {}).get("rows", [])
        income_9m: dict = {}
        if cum_rows:
            income_9m = self._extract_key_items(cum_rows, self._INCOME_KEYS)

        return {
            "period_description": period_desc,
            "headers": headers,
            "income": income,
            "income_9m": income_9m,   # cumulative 9M income (Q3 ZIPs only)
            "balance": balance,
            "cashflow": cashflow,
            "special_items_from_cf": special_items,
            "cashflow_cumulative": cashflow_cumulative,
            "unit": unit,
            "raw_parsed": parsed,  # keep full detail for XLSX tab
        }

    # ------------------------------------------------------------------
    # Full pipeline: symbol -> complete data
    # ------------------------------------------------------------------
    def fetch_full_data(self, symbol: str, use_cache: bool = True) -> dict:
        """
        Full pipeline to fetch and parse financial data for a symbol.

        1. Company profile
        2. Annual summary from company-highlight API (multiple years)
        3. Latest ZIP → extract quarterly summary from XLSX
        4. Load all previously cached quarterly summaries

        Returns dict with annual_data, quarterly_xlsx_data, xlsx_data, etc.
        """
        symbol = symbol.upper().strip()
        cache_file = CACHE_DIR / f"{symbol}_data.json"

        # Check cache (valid for 6 hours)
        if use_cache and cache_file.exists():
            age = time.time() - cache_file.stat().st_mtime
            if age < 6 * 3600:
                try:
                    with open(cache_file, "r") as f:
                        return json.load(f)
                except Exception:
                    pass

        result = {
            "symbol": symbol,
            "company": {},
            "annual_data": [],
            "quarterly_xlsx_data": [],  # structured quarterly data from ZIPs
            "factsheet_data": {},       # factsheet API: {income_statement, balance_sheet, cash_flow}
            "factsheet_ratios": [],     # factsheet financial ratios (multi-period)
            "latest_quarter": None,
            "xlsx_data": None,
            "zip_url": None,
            "error": None,
        }

        # 1. Company profile
        profile = self.get_company_profile_factsheet(symbol)
        profile_th = self._api_get(f"stock/{symbol}/profile", {"lang": "th"})
        if profile:
            result["company"] = {
                "name": profile.get("name", symbol),
                "name_th": profile_th.get("name", "") if profile_th else "",
                "sector": profile.get("sectorName", profile.get("sector", "")),
                "industry": profile.get("industryName", profile.get("industry", "")),
                "market": profile.get("market", "SET"),
                "fiscal_year_end": profile.get("fiscalYearEndDisplay", profile.get("fiscalYearEnd", "")),
            }
        else:
            result["company"] = {
                "name": profile_th.get("name", symbol) if profile_th else symbol,
                "name_th": profile_th.get("name", symbol) if profile_th else symbol,
                "sector": "N/A",
                "market": "SET",
            }

        # 2. Annual financial summary (multiple years)
        highlights = self.get_company_highlight_financial(symbol)
        if highlights:
            result["annual_data"] = [
                item for item in highlights
                if item.get("quarter", "") in ("Q9", "YE", "Q4", "")
            ] or highlights

        # 2b. Factsheet financial statements (multi-period, structured line items)
        #     Typically 4-5 periods: mix of quarterly + annual
        factsheet = {}
        for acct in ("income_statement", "balance_sheet", "cash_flow"):
            fs_data = self.get_factsheet_financialstatement(symbol, acct)
            if fs_data:
                factsheet[acct] = fs_data
        result["factsheet_data"] = factsheet

        # 2c. Factsheet financial ratios (multi-period)
        fs_ratios = self.get_factsheet_financial_ratio(symbol)
        if fs_ratios:
            result["factsheet_ratios"] = fs_ratios

        # 3. Download ALL historical financial statement ZIPs
        #    Uses news/search API with date ranges to find all FS filings,
        #    then downloads each ZIP and caches the parsed quarterly summary.
        cache_dir = self._quarterly_cache_dir(symbol)

        # 3a. Get latest ZIP info (for metadata display)
        zip_info = self.get_latest_fs_zip_url(symbol)
        if zip_info and zip_info.get("downloadUrl"):
            result["zip_url"] = zip_info["downloadUrl"]
            result["zip_info"] = {
                "quarter": zip_info.get("quarter"),
                "year": zip_info.get("year"),
                "status": zip_info.get("status"),
                "fsType": zip_info.get("fsTypeDescription"),
            }

        # 3b. Find all FS news items via news/search (with date ranges)
        fs_news = self.get_all_fs_news(symbol, years_back=7)

        # 3c. Download and parse each ZIP that isn't already cached
        for idx, news_item in enumerate(fs_news):
            headline = news_item.get("headline", "")
            quarter, year_ce = self._parse_fs_headline(headline)
            if not year_ce:
                continue

            # Skip if already cached
            cache_path = cache_dir / f"{year_ce}_{quarter}.json"
            if cache_path.exists():
                continue

            # Get download URL from news detail
            detail = self.get_fs_news_detail(str(news_item["id"]))
            if not detail or not detail.get("downloadUrl"):
                continue

            download_url = detail["downloadUrl"]

            try:
                zip_bytes = self.download_zip(download_url)
                if not zip_bytes:
                    continue
                xlsx_bytes = self.extract_xlsx_from_zip(zip_bytes)
                if not xlsx_bytes:
                    continue

                # For the first (latest) ZIP, keep full XLSX for detail tab
                if idx == 0:
                    result["xlsx_data"] = self.parse_financial_xlsx(xlsx_bytes)

                # Extract quarterly summary & persist
                q_summary = self.extract_quarterly_summary(xlsx_bytes)
                q_summary.pop("raw_parsed", None)
                q_summary["quarter"] = quarter
                q_summary["year"] = year_ce
                q_summary["download_url"] = download_url
                self._save_quarterly_cache(symbol, q_summary)
            except Exception:
                pass

        # 4. Load all cached quarterly summaries for this symbol
        result["quarterly_xlsx_data"] = self._load_all_quarterly_cache(symbol)

        # 4b. If some year will have FY but missing Q1/Q2 (so Strategy A cannot
        #     compute Q4), fetch that year's Q3 ZIP to obtain income_9m (9-month
        #     cumulative) so _add_q4_rows_from_xlsx can use Strategy B (Q4=FY-9M).
        #     The fetched Q3 is saved with "q4_helper": True so the display
        #     layer can exclude it from charts/tables.
        self._ensure_q3_helper_for_oldest_q4(symbol, fs_news, result["quarterly_xlsx_data"], cache_dir)

        # 4c. If a year still has FY + Q3 but no Q1/Q2 (Strategy B not possible
        #     because Q3 has no income_9m), fetch Q1 and Q2 of year+1 to get
        #     Q1/{year} and Q2/{year} from their prev columns. Saved as gap_filler.
        self._ensure_gap_filler_quarters(symbol, fs_news, result["quarterly_xlsx_data"], cache_dir)

        # Reload after potential helper/gap-filler downloads
        result["quarterly_xlsx_data"] = self._load_all_quarterly_cache(symbol)

        # 5. Latest quarter key financial data (for quick metrics)
        latest = self.get_key_financial_data(symbol)
        if latest:
            result["latest_quarter"] = latest

        # Cache full result (6 hours)
        try:
            with open(cache_file, "w") as f:
                json.dump(result, f, ensure_ascii=False, default=str)
        except Exception:
            pass

        return result

    # ------------------------------------------------------------------
    # Quarterly ZIP cache (persistent, survives 6-hour main cache)
    # ------------------------------------------------------------------
    def _quarterly_cache_dir(self, symbol: str) -> Path:
        d = CACHE_DIR / f"{symbol.upper()}_quarters"
        d.mkdir(exist_ok=True)
        return d

    def _save_quarterly_cache(self, symbol: str, q_summary: dict):
        """Save one quarter's XLSX summary to persistent cache."""
        q = q_summary.get("quarter", "")
        y = q_summary.get("year", "")
        if not q or not y:
            return
        cache_dir = self._quarterly_cache_dir(symbol)
        path = cache_dir / f"{y}_{q}.json"
        try:
            with open(path, "w") as f:
                json.dump(q_summary, f, ensure_ascii=False, default=str)
        except Exception:
            pass

    def _load_all_quarterly_cache(self, symbol: str) -> list[dict]:
        """Load all cached quarterly summaries, newest first."""
        cache_dir = self._quarterly_cache_dir(symbol)
        results = []
        for path in sorted(cache_dir.glob("*.json"), reverse=True):
            try:
                with open(path, "r") as f:
                    results.append(json.load(f))
            except Exception:
                pass
        return results

    def _ensure_q3_helper_for_oldest_q4(
        self,
        symbol: str,
        fs_news: list[dict],
        q_cache: list[dict],
        cache_dir: "Path",
    ) -> None:
        """
        Fetch Q3 helper ZIPs so that _add_q4_rows_from_xlsx can compute Q4
        via Strategy B (Q4 = FY − 9M cumulative) for years that will have FY
        but are missing Q1/Q2 in the final row set.

        Two scenarios trigger a helper fetch:
          1. Direct: a year in q_cache has FY (Q9/YE) but no Q3.
          2. Prev-column gap: a year Y has Q3 (and possibly FY) derived only
             from q_cache prev-columns but lacks Q1 and Q2 (because the years
             that would supply Q1/Q2 prev-columns are not in q_cache).

        The helper entry is used only for Q4 computation and must be excluded
        from all display tables/charts (financial_data.py filters it out via
        the q4_helper flag).
        """
        if not q_cache:
            return

        # Collect what years/quarters we already have directly in q_cache
        have: dict[int, set[str]] = {}
        for qd in q_cache:
            y = qd.get("year", 0)
            if isinstance(y, str):
                y = int(y) if str(y).isdigit() else 0
            q = qd.get("quarter", "")
            if y and q:
                have.setdefault(y, set()).add(q)

        # Determine which quarterly periods this company actually reports.
        # Some companies (e.g. banks) don't file Q2 at all — for those we
        # cannot fill the gap via a Q3 helper and must skip.
        company_reports_q2 = any(
            qd.get("quarter") == "Q2" for qd in q_cache
        )

        # Also compute what quarters each year will have AFTER prev-column
        # derivation.  If year Y has quarter Q in q_cache, then year Y-1 will
        # gain quarter Q via the prev-column mechanism in _build_quarterly.
        have_with_prev: dict[int, set[str]] = {y: set(qs) for y, qs in have.items()}
        for y, qs in have.items():
            prev_y = y - 1
            prev_derived: set[str] = set()
            if qs & {"Q9", "YE"}:
                prev_derived.add("FY")
            for q in ("Q1", "Q2", "Q3"):
                if q in qs:
                    prev_derived.add(q)
            if prev_derived:
                have_with_prev.setdefault(prev_y, set()).update(prev_derived)

        # Find years that will have FY but are missing Q1 and Q2 so that
        # Strategy A (FY - Q1 - Q2 - Q3) cannot compute Q4.
        # A Q3 helper ZIP provides income_9m (9-month cumulative) which
        # enables Strategy B (Q4 = FY - 9M).
        # Skip companies that never report Q2 (e.g. semi-annual filers) as a
        # helper fetch would not resolve the gap.
        needs_q3_helper: list[int] = []
        for y, qs in have_with_prev.items():
            # Check if this year has FY (either direct or via prev)
            has_fy = bool(qs & {"Q9", "YE", "FY"})
            if not has_fy:
                continue
            # If Q1, Q2, Q3 are all present → Strategy A will work, no helper needed
            if all(q in qs for q in ("Q1", "Q2", "Q3")):
                continue
            # If Q3 is missing entirely → original case (Strategy B requires helper)
            # If Q3 is present but Q1/Q2 missing AND company reports Q2 → need helper
            #   for income_9m so Strategy B (FY - 9M) can be applied.
            # If company never reports Q2 (e.g. banks) → skip, gap is structural.
            missing_q1_q2 = not ("Q1" in qs and "Q2" in qs)
            if "Q3" not in qs or (missing_q1_q2 and company_reports_q2):
                # Skip if a helper is already in q_cache (q4_helper=True)
                already_have_helper = any(
                    qd.get("year") == y and qd.get("quarter") == "Q3" and qd.get("q4_helper")
                    for qd in q_cache
                )
                if not already_have_helper:
                    needs_q3_helper.append(y)

        if not needs_q3_helper:
            return

        # Build a lookup: (quarter, year_ce) → news_item from fs_news
        news_index: dict[tuple[str, int], dict] = {}
        for item in fs_news:
            headline = item.get("headline", "")
            q, y = self._parse_fs_headline(headline)
            if y:
                news_index[(q, y)] = item

        for target_year in needs_q3_helper:
            helper_cache_path = cache_dir / f"{target_year}_Q3.json"
            if helper_cache_path.exists():
                # Already cached — ensure q4_helper flag is set
                try:
                    with open(helper_cache_path) as f:
                        data = json.load(f)
                    if not data.get("q4_helper"):
                        data["q4_helper"] = True
                        with open(helper_cache_path, "w") as f:
                            json.dump(data, f, ensure_ascii=False, default=str)
                except Exception:
                    pass
                continue

            news_item = news_index.get(("Q3", target_year))
            if not news_item:
                continue

            try:
                detail = self.get_fs_news_detail(str(news_item["id"]))
                if not detail or not detail.get("downloadUrl"):
                    continue
                download_url = detail["downloadUrl"]
                zip_bytes = self.download_zip(download_url)
                if not zip_bytes:
                    continue
                xlsx_bytes = self.extract_xlsx_from_zip(zip_bytes)
                if not xlsx_bytes:
                    continue

                q_summary = self.extract_quarterly_summary(xlsx_bytes)
                q_summary.pop("raw_parsed", None)
                q_summary["quarter"] = "Q3"
                q_summary["year"] = target_year
                q_summary["download_url"] = download_url
                q_summary["q4_helper"] = True   # ← mark as helper-only

                with open(helper_cache_path, "w") as f:
                    json.dump(q_summary, f, ensure_ascii=False, default=str)
            except Exception:
                pass

    def _ensure_gap_filler_quarters(
        self,
        symbol: str,
        fs_news: list[dict],
        q_cache: list[dict],
        cache_dir: "Path",
    ) -> None:
        """
        For years that will have FY + Q3 but are missing Q1 and Q2 (so neither
        Strategy A nor Strategy B can compute Q4), fetch the Q1 and Q2 ZIPs of
        the *following* year (year+1).  Their prev columns provide Q1/{year} and
        Q2/{year}, enabling Strategy A: Q4 = FY - (Q1+Q2+Q3).

        Saved with "gap_filler": True so _build_quarterly knows they are
        supplementary and should not be shown as display rows themselves (only
        their prev columns are used).

        This handles the common case where the oldest cached quarterly data is
        Q3/{year} but Q1 and Q2 of that same year are absent because the API
        only exposes news back to that Q3.
        """
        if not q_cache:
            return

        # Collect what years/quarters we have in q_cache (direct entries only)
        have: dict[int, set[str]] = {}
        for qd in q_cache:
            y = qd.get("year", 0)
            if isinstance(y, str):
                y = int(y) if str(y).isdigit() else 0
            q = qd.get("quarter", "")
            if y and q:
                have.setdefault(y, set()).add(q)

        # Compute have_with_prev (same logic as _ensure_q3_helper_for_oldest_q4)
        company_reports_q2 = any(qd.get("quarter") == "Q2" for qd in q_cache)
        if not company_reports_q2:
            return  # company doesn't report Q2 — gap-filler won't help

        have_with_prev: dict[int, set[str]] = {y: set(qs) for y, qs in have.items()}
        for y, qs in have.items():
            prev_y = y - 1
            prev_derived: set[str] = set()
            if qs & {"Q9", "YE"}:
                prev_derived.add("FY")
            for q in ("Q1", "Q2", "Q3"):
                if q in qs:
                    prev_derived.add(q)
            if prev_derived:
                have_with_prev.setdefault(prev_y, set()).update(prev_derived)

        # Find years that will have FY + Q3 but not Q1 or Q2 after prev derivation
        needs_gap_fill: list[int] = []
        for y, qs in have_with_prev.items():
            has_fy = bool(qs & {"Q9", "YE", "FY"})
            if not has_fy:
                continue
            if all(q in qs for q in ("Q1", "Q2", "Q3")):
                continue  # Strategy A already works
            if "Q3" not in qs:
                continue  # no Q3 → handled by _ensure_q3_helper_for_oldest_q4
            # Has FY + Q3 but missing Q1 or Q2
            missing = [q for q in ("Q1", "Q2") if q not in qs]
            if not missing:
                continue
            # Check if gap_filler entries already exist in q_cache for year+1
            next_year = y + 1
            already_filled = all(
                any(
                    qd.get("year") == next_year
                    and qd.get("quarter") == q
                    and qd.get("gap_filler")
                    for qd in q_cache
                )
                for q in missing
            )
            if not already_filled:
                needs_gap_fill.append(y)

        if not needs_gap_fill:
            return

        # Build news index: (quarter, year_ce) → news_item
        news_index: dict[tuple[str, int], dict] = {}
        for item in fs_news:
            headline = item.get("headline", "")
            q, y = self._parse_fs_headline(headline)
            if y:
                news_index[(q, y)] = item

        for target_year in needs_gap_fill:
            next_year = target_year + 1
            # Determine which quarters of next_year to fetch as gap-fillers
            qs_in_prev = have_with_prev.get(target_year, set())
            quarters_to_fetch = [q for q in ("Q1", "Q2") if q not in qs_in_prev]

            for quarter in quarters_to_fetch:
                gap_cache_path = cache_dir / f"{next_year}_{quarter}_gap.json"
                if gap_cache_path.exists():
                    # Already cached — ensure gap_filler flag is set
                    try:
                        with open(gap_cache_path) as f:
                            data = json.load(f)
                        if not data.get("gap_filler"):
                            data["gap_filler"] = True
                            with open(gap_cache_path, "w") as f:
                                json.dump(data, f, ensure_ascii=False, default=str)
                    except Exception:
                        pass
                    continue

                news_item = news_index.get((quarter, next_year))
                if not news_item:
                    continue

                try:
                    detail = self.get_fs_news_detail(str(news_item["id"]))
                    if not detail or not detail.get("downloadUrl"):
                        continue
                    download_url = detail["downloadUrl"]
                    zip_bytes = self.download_zip(download_url)
                    if not zip_bytes:
                        continue
                    xlsx_bytes = self.extract_xlsx_from_zip(zip_bytes)
                    if not xlsx_bytes:
                        continue

                    q_summary = self.extract_quarterly_summary(xlsx_bytes)
                    q_summary.pop("raw_parsed", None)
                    q_summary["quarter"] = quarter
                    q_summary["year"] = next_year
                    q_summary["download_url"] = download_url
                    q_summary["gap_filler"] = True  # ← mark as gap-filler only

                    with open(gap_cache_path, "w") as f:
                        json.dump(q_summary, f, ensure_ascii=False, default=str)
                except Exception:
                    pass
